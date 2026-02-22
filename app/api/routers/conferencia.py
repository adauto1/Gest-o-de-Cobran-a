import io
import json
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.web import render, require_login
from app.models import ConferenciaTitulos, ReportSnapshot
from app.services.conferencia_inteligente_service import process_smart_reconciliation

router = APIRouter()


@router.get("/conferencia", response_class=HTMLResponse)
def conferencia_page(request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    if user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Apenas administradores podem acessar esta página.")

    last_process = db.query(ConferenciaTitulos).order_by(ConferenciaTitulos.data_processamento.desc()).first()
    detalhes = []
    resumo = None
    msg = None
    if last_process:
        detalhes = json.loads(last_process.detalhes_json)
        resumo = json.loads(last_process.resumo_json)
        # Garantir que todas as chaves existam
        for key in ["normal_qtd", "normal_valor", "divergencia_qtd", "divergencia_valor",
                    "suspeita_qtd", "suspeita_valor", "total_analisado"]:
            resumo.setdefault(key, 0)
        resumo.setdefault("diagnostico", [])
        resumo.setdefault("snap_a_rec_prev_str", "-")
        resumo.setdefault("snap_a_rec_curr_str", "-")
        resumo.setdefault("snap_rec_str", "-")
        resumo["data"] = last_process.data_processamento.strftime("%d/%m/%Y %H:%M")

    return render("conferencia.html", request=request, user=user, title="Conferência de Títulos",
                  active_page="conferencia", detalhes=detalhes, resumo=resumo, msg=msg)


@router.post("/api/conferencia/processar")
async def processar_conferencia(
    request: Request,
    file_recebido: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    user = require_login(request, db)
    if user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Apenas administradores.")
    content_recebido = await file_recebido.read() if file_recebido else None
    results = process_smart_reconciliation(db, content_recebido)
    
    if "error" in results:
        raise HTTPException(status_code=400, detail=results["error"])
        
    return results


@router.get("/api/conferencia/exportar-excel")
def exportar_excel(request: Request, db: Session = Depends(get_db)):
    require_login(request, db)

    last = db.query(ConferenciaTitulos).order_by(ConferenciaTitulos.data_processamento.desc()).first()
    if not last:
        raise HTTPException(status_code=404, detail="Nenhum relatório de conferência encontrado.")

    detalhes = json.loads(last.detalhes_json)
    resumo = json.loads(last.resumo_json)
    data_proc = last.data_processamento.strftime("%d/%m/%Y %H:%M")

    wb = Workbook()
    ws = wb.active
    ws.title = "Conferência de Títulos"

    # Estilos
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    fill_green = PatternFill("solid", fgColor="DCFCE7")
    fill_yellow = PatternFill("solid", fgColor="FEF3C7")
    fill_red = PatternFill("solid", fgColor="FEE2E2")

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Conferência de Títulos — Processado em {data_proc}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Resumo
    ws["A2"] = f"✅ Quitadas Normalmente: {resumo.get('normal_qtd', 0)}"
    ws["C2"] = f"⚠️ Divergência de Valor: {resumo.get('divergencia_qtd', 0)}"
    ws["E2"] = f"🔴 Suspeitas de Exclusão: {resumo.get('suspeita_qtd', 0)}"
    ws["A3"] = ""

    # Cabeçalho da tabela
    headers = ["Cliente", "Pedido", "Vencimento", "Valor ERP (R$)", " Situação", "Evidência", "Snap. Anterior", "Snap. Atual"]
    ws.append(headers)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for item in detalhes:
        grupo = item.get("grupo", "")
        fill = fill_green if grupo == "BAIXA JUSTIFICADA" else fill_yellow if grupo in ("DIVERGENCIA", "PARCELA REMOVIDA") else fill_red
        row = [
            item.get("cliente", ""),
            item.get("pedido", ""),
            item.get("venc", ""),
            item.get("valor", 0),
            item.get("situacao", ""),
            item.get("evidencia", ""),
            item.get("snapshot_ant", ""),
            item.get("snapshot_atu", ""),
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # Largura das colunas
    col_widths = [35, 15, 14, 16, 25, 20, 20, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Formatar colunas de valor como moeda
    for row in ws.iter_rows(min_row=hdr_row + 1, min_col=4, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"conferencia_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/api/conferencia/zerar")
async def zerar_conferencia(request: Request, db: Session = Depends(get_db)):
    user = require_login(request, db)
    if user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Apenas administradores podem zerar o relatório.")
    db.query(ConferenciaTitulos).delete()
    db.query(ReportSnapshot).delete()
    db.commit()
    return {"success": True}
